import csv
import os
import re
import sys
from datetime import datetime
import shutil
import subprocess
import platform
import time
from email_store import EmailStore
try:
    from openpyxl.worksheet.pagebreak import PageBreak  # openpyxl >= 3.1
except Exception:
    PageBreak = None

# Optional heavy dependencies may not be installed in the runtime used for quick GUI previews.
# Import them conditionally and fall back so the preview/detection features remain usable.
HAS_OPENPYXL = False
HAS_REPORTLAB = False
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    HAS_OPENPYXL = True
except Exception:
    load_workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except Exception:
    # ReportLab not available; PDF generation will be disabled but detection/preview still work.
    colors = None
    SimpleDocTemplate = None
    Paragraph = None
    Spacer = None
    Table = None
    TableStyle = None
    getSampleStyleSheet = None
    ParagraphStyle = None
    inch = None
    pdfmetrics = None
    TTFont = None


PERIOD_HEADER_REGEX = re.compile(r'([A-Za-z]{3,}\.?)\s*\d{0,2}\s*[-–—]\s*\d{1,2}', re.IGNORECASE)


def _normalize_header_text(header: str) -> str:
    """Trim, collapse whitespace, and normalise dashes for header text."""
    header = (header or "").replace("\u2013", "-").replace("\u2014", "-")
    header = re.sub(r"\s+", " ", header.strip())
    return header


def _is_period_header(header: str) -> bool:
    cleaned = _normalize_header_text(header)
    return bool(cleaned and PERIOD_HEADER_REGEX.search(cleaned))

class DynamicPayrollProcessor:
    def __init__(self):
        self.periods = []
        self.employees = []
        # Base dir: when frozen by PyInstaller, resources live in sys._MEIPASS
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            self.base_dir = os.path.abspath(sys._MEIPASS)
        else:
            self.base_dir = os.path.abspath(os.getcwd())

        # User-writable application data directory for outputs/logs/templates editable by user
        homedir = os.path.expanduser('~')
        if os.name == 'nt':
            appdata = os.getenv('LOCALAPPDATA') or os.path.join(homedir, 'AppData', 'Local')
        else:
            appdata = os.getenv('XDG_DATA_HOME') or os.path.join(homedir, '.local', 'share')
        self.user_data_dir = os.path.join(appdata, 'dynamic_payroll')

        self.template_path = self._resolve_template_path()

        # Create organized folder structure (under user_data_dir)
        self.create_folders()

        # Persistent email store
        try:
            store_path = os.path.join(self.user_data_dir, "email_addresses.sqlite3")
            self.email_store = EmailStore(store_path)
            # merge existing CSV (if present) for backward compatibility
            legacy_csv = os.path.join(os.getcwd(), "emails.csv")
            if os.path.exists(legacy_csv):
                self.email_store.import_from_csv(legacy_csv)
        except Exception:
            # fallback to a dummy store to avoid breaking execution
            self.email_store = None

    def create_folders(self):
        """Create organized folder structure"""
        folders = [
            os.path.join(self.user_data_dir, "input_csv"),
            os.path.join(self.user_data_dir, "templates"),
            os.path.join(self.user_data_dir, "output_pdfs"),
            os.path.join(self.user_data_dir, "logs"),
        ]

        for folder in folders:
            try:
                if not os.path.exists(folder):
                    os.makedirs(folder, exist_ok=True)
            except Exception:
                # best-effort; don't crash if user dir cannot be created
                pass

    def detect_periods(self, csv_filename):
        """Detect period columns dynamically from CSV headers"""
        self.periods = []
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

        for encoding in encodings_to_try:
            try:
                with open(csv_filename, 'r', encoding=encoding, newline='') as fh:
                    reader = csv.reader(fh)
                    rows = list(reader)

                    # Find header row: prefer the row that contains 'per hour' (explicit header).
                    # If not found, look for a row that contains at least two period-like headers
                    header_row_idx = None
                    # First pass: look for explicit 'per hour' marker
                    for i, row in enumerate(rows):
                        joined = ' '.join([c.lower() for c in row if c])
                        if 'per hour' in joined:
                            header_row_idx = i
                            break
                    # Second pass: fallback - find a row with multiple period-like entries
                    if header_row_idx is None:
                        for i, row in enumerate(rows):
                            period_matches = sum(1 for c in row if c and re.match(period_pattern, c.strip()))
                            if period_matches >= 2 and len(row) >= 4:
                                header_row_idx = i
                                break

                    if header_row_idx is None:
                        continue

                    # Some CSVs (like the provided Payroll.csv) use a two-line header:
                    # one row with broad labels (Seq., NAME, RATE, etc.) and the next
                    # row contains subheaders such as 'per hour' and the period dates.
                    # If the row above the detected header has meaningful labels, merge
                    # them so we get full header names per column.
                    headers = []
                    prev_row = rows[header_row_idx - 1] if header_row_idx > 0 else None
                    cur_row = rows[header_row_idx]
                    max_cols = max(len(prev_row) if prev_row is not None else 0, len(cur_row))
                    for i in range(max_cols):
                        parts = []
                        if prev_row is not None and i < len(prev_row) and prev_row[i]:
                            parts.append(str(prev_row[i]).strip())
                        if i < len(cur_row) and cur_row[i]:
                            parts.append(str(cur_row[i]).strip())
                        header = _normalize_header_text(' '.join(parts))
                        headers.append(header)

                    # Save merged headers for external consumers (GUI) so headings mirror the CSV
                    try:
                        self.merged_headers = headers
                    except Exception:
                        self.merged_headers = []

                    # Detect period columns by regex on headers
                    for i, header in enumerate(headers):
                        if _is_period_header(header):
                            match = PERIOD_HEADER_REGEX.search(header or "")
                            short_name = _normalize_header_text(match.group(0)) if match else _normalize_header_text(header)
                            self.periods.append({
                                'name': short_name,
                                'display_name': _normalize_header_text(header),
                                'column_index': i,
                                'hours_col': i,
                                'amount_col': None
                            })

                    if self.periods:
                        print(f"Successfully loaded CSV with encoding: {encoding}")
                        print(f"Detected {len(self.periods)} pay periods:")
                        for period in self.periods:
                            print(f"  - {period['name']}")
                        return True

            except UnicodeDecodeError:
                continue

        print("Could not find header row with 'per hour' or encoding issues")
        return False
    def _resolve_template_path(self):
        """
        Find the Excel template in common locations or via env var.
        Returns absolute path or None.
        """
        # Candidate locations (in order): env override, bundled base_dir/templates, user_data_dir/templates, repo templates
        env_path = (os.environ.get("PAYSLIP_TEMPLATE") or "").strip() or None
        candidates = []
        if env_path:
            candidates.append(env_path)
        # bundled template inside base_dir (when frozen)
        candidates.append(os.path.join(self.base_dir, 'templates', 'Template Payslip.xlsx'))
        # user writable templates folder
        candidates.append(os.path.join(self.user_data_dir, 'templates', 'Template Payslip.xlsx'))
        # repo-local templates (development)
        candidates.append(os.path.join(os.getcwd(), 'templates', 'Template Payslip.xlsx'))
        candidates.append(os.path.join(os.getcwd(), 'Template Payslip.xlsx'))

        for c in candidates:
            try:
                if c and os.path.exists(c):
                    return os.path.abspath(c)
            except Exception:
                continue
        return None

    # -- CSV numeric parsing helpers --
    def safe_float(self, s, default=0.0):
        """Convert string to float robustly: strip commas, parentheses, stray characters; return default on failure."""
        if s is None:
            return default
        s = str(s).strip()
        if not s or s in ['-', '—', '–']:
            return default
        # Remove commas and currency symbols
        s = s.replace(',', '').replace('$', '')
        # Replace multiple spaces or stray non-numeric chars except dot and minus
        import re
        s = re.sub(r"[^0-9.\-]", '', s)
        try:
            return float(s) if s not in ['', '.', '-'] else default
        except ValueError:
            return default

    def parse_percent(self, s, default=0.0):
        """Parse percent-like strings (e.g., '12%', '12.5') into a decimal fraction (0.12)."""
        if s is None:
            return default
        s = str(s).strip()
        if not s or s in ['-', '—', '–']:
            return default
        # Remove percent sign and whitespace
        s = s.replace('%', '').strip()
        val = self.safe_float(s, default=None)
        if val is None:
            return default
        return val / 100.0

    def load_employee_data(self, csv_filename):
        """Load employee data with dynamic period detection"""
        self.employees = []
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

        for encoding in encodings_to_try:
            try:
                with open(csv_filename, 'r', encoding=encoding, newline='') as fh:
                    reader = csv.reader(fh)
                    rows = list(reader)

            # Find header row (prefer the one that mentions "per hour")
                header_row_idx = None
                for i, row in enumerate(rows):
                    joined = ' '.join([c.lower() for c in row if c])
                    if 'per hour' in joined:
                        header_row_idx = i
                        break
                if header_row_idx is None:
                    # fallback: row that looks like it has multiple period headers
                    for i, row in enumerate(rows):
                        period_matches = sum(1 for c in row if c and re.match(period_pattern, c.strip()))
                    if period_matches >= 2 and len(row) >= 4:
                            header_row_idx = i
                            break
                if header_row_idx is None:
                # try next encoding
                    continue

                # Merge possible two-line headers (previous row + current row)
                prev_row = rows[header_row_idx - 1] if header_row_idx > 0 else None
                cur_row = rows[header_row_idx]
                headers = []
                max_cols = max(len(prev_row) if prev_row is not None else 0, len(cur_row))
                for j in range(max_cols):
                    parts = []
                    if prev_row is not None and j < len(prev_row) and prev_row[j]:
                        parts.append(str(prev_row[j]).strip())
                    if j < len(cur_row) and cur_row[j]:
                        parts.append(str(cur_row[j]).strip())
                    headers.append(_normalize_header_text(' '.join(parts)))

                # --- header maps (order-agnostic) ---
                header_map_lower = {h.lower(): idx for idx, h in enumerate(headers)}

                def _norm(s: str) -> str:
                    return re.sub(r'[^a-z0-9]+', '', (s or '').lower())

                header_map_norm = {_norm(h): i for i, h in enumerate(headers)}

                def find_col(*parts):
                    """Return index of the first header that contains ALL normalized parts."""
                    parts_n = [_norm(p) for p in parts]
                    for h_norm, idx in header_map_norm.items():
                        if all(p in h_norm for p in parts_n):
                            return idx
                    return None

                # Detect periods from headers if not already available
                if not self.periods:
                    self.periods = []
                    for j, h in enumerate(headers):
                        if _is_period_header(h):
                            self.periods.append({'name': _normalize_header_text(h), 'column_index': j, 'hours_col': j, 'amount_col': None})

                # Heuristics for key columns
                seq_idx = next((idx for h, idx in header_map_lower.items() if 'seq' in h or 'sequence' in h), None)
                name_idx = next((idx for h, idx in header_map_lower.items() if 'name' in h), None)
                rate_idx = next((idx for h, idx in header_map_lower.items() if ('per hour' in h) or ('hour' in h) or ('rate' in h)), None)

                # Fallbacks
                if seq_idx is None:
                    seq_idx = 0
                if name_idx is None:
                    name_idx = 2 if len(headers) > 2 else 1
                # rate_idx may be None → we’ll treat hourly_rate as 0 in that case

                data_rows = rows[header_row_idx + 1:]

                idx_email = find_col('email')
                idx_work_email = find_col('work', 'email')
                idx_personal_email = find_col('personal', 'email')

                for row in data_rows:
                    if not row:
                        continue
                    # ensure this looks like a data row
                    if seq_idx >= len(row) or not str(row[seq_idx]).strip().isdigit():
                        continue

                    try:
                        employee = {
                            'seq': row[seq_idx].strip(),
                            'account_no': row[1].strip() if len(row) > 1 else '',
                            'name': row[name_idx].strip() if name_idx is not None and name_idx < len(row) else '',
                            'hourly_rate': self.safe_float(row[rate_idx]) if rate_idx is not None and rate_idx < len(row) else 0,
                            'email': '',
                            'work_email': '',
                            'personal_email': '',
                            'periods': {},
                            'total_gross': 0,
                            'tax_rate': 0,
                            'withholding_rate': 0,
                            'withholding_amount': 0,
                            'p_tax_rate': 0,
                            'percent_tax': 0,
                            'adjustment_hours': 0,
                            'adjustment_amount': 0,
                            'total_tax_deductions': 0,
                            'net_amount_received': 0
                        }

                        # Extract email columns when present
                        if idx_email is not None and idx_email < len(row) and row[idx_email].strip():
                            employee['email'] = row[idx_email].strip()
                        if idx_work_email is not None and idx_work_email < len(row) and row[idx_work_email].strip():
                            employee['work_email'] = row[idx_work_email].strip()
                        if idx_personal_email is not None and idx_personal_email < len(row) and row[idx_personal_email].strip():
                            employee['personal_email'] = row[idx_personal_email].strip()

                        # Extract period hours and compute amount per period
                        for period in self.periods:
                            col_idx = period['column_index']
                            if col_idx < len(row):
                                raw_val = (row[col_idx] or '').strip()
                                hours = self.safe_float(raw_val, default=0)
                                rate = employee.get('hourly_rate', 0) or 0
                                amount = hours * rate if rate else 0
                                period_data = {'hours': hours, 'amount': amount, 'raw': raw_val}
                                canonical_name = period.get('name') or _normalize_header_text(headers[col_idx] if col_idx < len(headers) else '')
                                employee['periods'][canonical_name] = period_data

                                display_name = period.get('display_name')
                                if display_name and display_name != canonical_name:
                                    employee['periods'][display_name] = period_data

                                employee['total_gross'] += amount

                        # --- robust summary/tax mapping by header names ---
                        # --- robust summary/tax mapping by header names ---
                                idx_amt_earned = find_col('amount', 'earned')            # "AMOUNT EARNED"
                                idx_net_amt    = find_col('net', 'amount', 'earned')     # "NET AMOUNT EARNED" (do NOT overwrite gross)
                                idx_tax_rate   = find_col('w', 'tax', 'rate')            # "W/ TAX RATE"
                                idx_wh_tax     = find_col('holding', 'tax')              # "W/HOLDING TAX"
                                idx_p_rate     = find_col('p', 'tax', 'rate')            # "P-TAX RATE"
                                idx_percent    = find_col('percent', 'tax')              # "PERCENT. TAX"
                                idx_total_ded  = find_col('total', 'tax', 'deduct')      # "TOTAL TAX DEDUCTIONS"
                                idx_net_recv   = find_col('net', 'amount', 'received')   # "NET AMOUNT RECEIVED"

                        # --- adjustment columns (hours/amount) ---
                                idx_adj_hours  = find_col('adjust', 'hour')
                                idx_adj_amount = (find_col('adjust', 'amount')
                                    or find_col('adjust', 'add', 'tax')
                                    or find_col('addl', 'tax')
                                    or find_col('adjustment', 'tax'))

                            # set adjustment hours
                                if idx_adj_hours is not None and idx_adj_hours < len(row) and row[idx_adj_hours].strip():
                                    employee['adjustment_hours'] = self.safe_float(row[idx_adj_hours], default=0)

                            # prefer explicit amount if present
                                adj_amount_from_csv = None
                                if idx_adj_amount is not None and idx_adj_amount < len(row) and row[idx_adj_amount].strip():
                                    adj_amount_from_csv = self.safe_float(row[idx_adj_amount], default=0)

                            # AMOUNT EARNED can override the computed gross (sum of periods)
                                if idx_amt_earned is not None and idx_amt_earned < len(row) and row[idx_amt_earned].strip():
                                    employee['total_gross'] = self.safe_float(row[idx_amt_earned], default=employee.get('total_gross', 0))

                        # NET AMOUNT EARNED should NOT overwrite gross; store separately
                                net_amt_from_csv = None
                                if idx_net_amt is not None and idx_net_amt < len(row) and row[idx_net_amt].strip():
                                    net_amt_from_csv = self.safe_float(row[idx_net_amt], default=None)
                                    employee['net_amount_earned'] = net_amt_from_csv

                        # If CSV didn't give adjustment amount, compute it from hours × rate
                                calc_adj = (employee.get('adjustment_hours', 0) or 0) * (employee.get('hourly_rate', 0) or 0)
                                if adj_amount_from_csv is not None:
                                    employee['adjustment_amount'] = adj_amount_from_csv
                                elif calc_adj:
                                    employee['adjustment_amount'] = calc_adj
                                else:
                                    employee['adjustment_amount'] = 0

                                # If CSV provided NET AMOUNT EARNED but not an adjustment amount,
                                # infer adjustment = gross - net (when sensible)
                                if (not adj_amount_from_csv) and (net_amt_from_csv is not None):
                                    gross = float(employee.get('total_gross', 0) or 0)
                                    diff = gross - net_amt_from_csv
                                # accept small negatives as zero (rounding)
                                    if abs(diff) > 1e-6:
                                        employee['adjustment_amount'] = diff

                                # tax/withholding/percent/total deductions/net received
                                if idx_tax_rate is not None and idx_tax_rate < len(row) and row[idx_tax_rate].strip():
                                    employee['tax_rate'] = self.parse_percent(row[idx_tax_rate], default=employee.get('tax_rate', 0))

                                if idx_wh_tax is not None and idx_wh_tax < len(row) and row[idx_wh_tax].strip():
                                    raw_wh = row[idx_wh_tax].strip()
                                    if '%' in raw_wh:
                                        employee['withholding_rate'] = self.parse_percent(raw_wh, default=employee.get('withholding_rate', 0))
                                    else:
                                        employee['withholding_amount'] = self.safe_float(raw_wh, default=employee.get('withholding_amount', 0))

                                if idx_p_rate is not None and idx_p_rate < len(row) and row[idx_p_rate].strip():
                                    employee['p_tax_rate'] = self.parse_percent(row[idx_p_rate], default=employee.get('p_tax_rate', 0))

                                if idx_percent is not None and idx_percent < len(row) and row[idx_percent].strip():
                                    employee['percent_tax'] = self.safe_float(row[idx_percent], default=employee.get('percent_tax', 0))

                                if idx_total_ded is not None and idx_total_ded < len(row) and row[idx_total_ded].strip():
                                    employee['total_tax_deductions'] = self.safe_float(row[idx_total_ded], default=employee.get('total_tax_deductions', 0))

                                if idx_net_recv is not None and idx_net_recv < len(row) and row[idx_net_recv].strip():
                                    employee['net_amount_received'] = self.safe_float(row[idx_net_recv], default=employee.get('net_amount_received', 0))

                                primary_email = next(
                                    (val for val in (employee.get('email'), employee.get('work_email'), employee.get('personal_email')) if val),
                                    None
                                )
                                if primary_email:
                                    employee['email'] = primary_email
                                    try:
                                        self.remember_employee_email(employee, primary_email)
                                    except Exception:
                                        pass

                        self.employees.append(employee)

                    except Exception as e:
                        print(f"Skipping row due to error: {e}")
                        continue

                if self.employees:
                    print(f"Successfully loaded CSV with encoding: {encoding}")
                    print(f"Loaded {len(self.employees)} employees successfully")
                    self._attach_persistent_emails()
                    return True

            except UnicodeDecodeError:
                # try next encoding
                continue
            except Exception as e:
                print(f"Error reading CSV with encoding {encoding}: {e}")
                continue

        print("Could not load employee data - encoding issues or no data found")
        return False

    def _attach_persistent_emails(self):
        if not getattr(self, 'email_store', None):
            return
        try:
            self.email_store.apply_to_employees(self.employees)
        except Exception:
            pass

    def remember_employee_email(self, employee, email):
        if not getattr(self, 'email_store', None):
            return
        try:
            self.email_store.remember_from_employee(employee, email)
        except Exception:
            pass


    def calculate_employee_payroll(self, employee):
        """Compute totals consistent with the Excel template."""
        gross = float(employee.get('total_gross', 0) or 0)              # Salary Earned
        adj   = float(employee.get('adjustment_amount', 0) or 0)        # D14: Adjustment (Add'l TAX)
        sub_total = gross - adj                                         # D15

        # Deductions are computed from gross (matches the template’s right block)
        w_tax       = gross * float(employee.get('tax_rate', 0) or 0)
        withholding = gross * float(employee.get('withholding_rate', 0) or 0)
        p_tax       = gross * float(employee.get('p_tax_rate', 0) or 0)

        total_deductions = w_tax + withholding + p_tax                  # I17
        net_pay = sub_total - total_deductions                          # D17

        return {
            'gross_pay': gross,
            'sub_total': sub_total,
            'w_tax': w_tax,
            'withholding': withholding,
            'p_tax': p_tax,
            'total_deductions': total_deductions,
            'net_pay': net_pay
        }

    def generate_pdf_payslip(self, employee, output_filename, withholding_placement='both'):
        """Generate PDF payslip; prefer Excel template, else fallback to ReportLab."""
        if HAS_OPENPYXL and self.template_path:
            print(f"[template] Using: {self.template_path}")
            ok = self.generate_pdf_from_excel_template(employee, output_filename, withholding_placement)
            if ok:
                return True
            print("[template] Export failed; falling back to ReportLab.")
        else:
            if not HAS_OPENPYXL:
                print("[template] openpyxl not available; falling back.")
            if not self.template_path:
                print("[template] Excel template not found; falling back.")
        # If ReportLab wasn't available at import time, try to import now (handles runtime installs)
        global HAS_REPORTLAB, colors, SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, getSampleStyleSheet, ParagraphStyle, inch, pdfmetrics, TTFont, A4, letter
        if not HAS_REPORTLAB:
            try:
                import reportlab.lib.colors as _colors
                import reportlab.lib.pagesizes as _pagesizes
                import reportlab.lib.styles as _styles
                import reportlab.lib.units as _units
                import reportlab.platypus as _platypus
                import reportlab.pdfbase as _pdfbase
                import reportlab.pdfbase.ttfonts as _ttf
                colors = _colors
                A4 = _pagesizes.A4
                letter = _pagesizes.letter
                getSampleStyleSheet = _styles.getSampleStyleSheet
                ParagraphStyle = _styles.ParagraphStyle
                inch = _units.inch
                SimpleDocTemplate = _platypus.SimpleDocTemplate
                Paragraph = _platypus.Paragraph
                Spacer = _platypus.Spacer
                Table = _platypus.Table
                TableStyle = _platypus.TableStyle
                pdfmetrics = _pdfbase
                TTFont = _ttf.TTFont
                HAS_REPORTLAB = True
            except Exception:
                print('ReportLab not installed - skipping PDF generation')
                return False

        try:
            # Create PDF document
            doc = SimpleDocTemplate(output_filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )

            story.append(Paragraph("PAMANTASAN NG LUNGSOD NG VALENZUELA", title_style))
            story.append(Paragraph("PART-TIME FACULTY PAYSLIP", title_style))
            story.append(Spacer(1, 20))

            # Pay period and date (include all detected periods in breakdown header)
            current_date = datetime.now().strftime("%B %d, %Y")
            period_names = list(employee.get('periods', {}).keys())
            period_text = f"Pay Periods: {', '.join(period_names)} | Generated: {current_date}"

            period_style = ParagraphStyle(
                'PeriodStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1
            )
            story.append(Paragraph(period_text, period_style))
            story.append(Spacer(1, 20))

            # Employee Information
            emp_info_style = ParagraphStyle(
                'EmpInfo',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=10
            )

            story.append(Paragraph("<b>EMPLOYEE INFORMATION</b>", emp_info_style))
            story.append(Paragraph(f"Sequence No.: {employee['seq']}", emp_info_style))
            if employee['account_no']:
                story.append(Paragraph(f"Account No.: {employee['account_no']}", emp_info_style))
            story.append(Paragraph(f"Name: {employee['name']}", emp_info_style))
            story.append(Spacer(1, 15))

            # Pay Details Table
            story.append(Paragraph("<b>PAY DETAILS</b>", emp_info_style))

            # Create period breakdown table: each detected period gets its own row
            period_data = [['DATE', 'HOURS EARNED', 'RATE', 'SALARY EARNED']]
            total_hours = 0.0
            total_salary = 0.0

            for period_name, period_info in employee.get('periods', {}).items():
                hours = float(period_info.get('hours', 0) or 0)
                rate = float(employee.get('hourly_rate', 0) or 0)
                salary = hours * rate
                period_data.append([
                    period_name,
                    f"{hours:.2f}",
                    f"{rate:.2f}",
                    f"{salary:.2f}"
                ])
                total_hours += hours
                total_salary += salary

            # Totals row
            period_data.append(['TOTAL', f"{total_hours:.2f}", '', f"{total_salary:.2f}"])

            period_table = Table(period_data, colWidths=[200, 80, 80, 100])
            period_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(period_table)
            story.append(Spacer(1, 20))

            # Payroll Calculations
            payroll = self.calculate_employee_payroll(employee)

            story.append(Paragraph("<b>PAYROLL SUMMARY</b>", emp_info_style))

            # Determine withholding/percentage tax placement for 15th and 30th
            # If user selected '15' or '30', place totals accordingly; if 'both', split proportionally
            w_rate = float(employee.get('withholding_rate', 0) or 0)
            p_rate = float(employee.get('p_tax_rate', 0) or 0)

            # Compute base amounts from totals or derived salary
            gross = float(employee.get('total_gross', total_salary) or total_salary)

            # Prefer absolute amounts from CSV when present; otherwise compute from rates
            w_amount = float(employee.get('withholding_amount', 0) or 0)
            if w_amount == 0 and w_rate:
                w_amount = gross * w_rate

            p_amount = float(employee.get('percent_tax', 0) or 0)
            if p_amount == 0 and p_rate:
                p_amount = gross * p_rate

            wp = withholding_placement or 'both'

            def split_amount(amount, place):
                if place == 'both':
                    return amount / 2.0, amount / 2.0
                if place == '15':
                    return amount, 0.0
                return 0.0, amount

            w_15, w_30 = split_amount(w_amount, wp)
            p_15, p_30 = split_amount(p_amount, wp)

            total_15 = w_15 + p_15
            total_30 = w_30 + p_30
            total_deductions = total_15 + total_30
            net_pay = gross - total_deductions

            # Build labels: always show percentage when a rate exists.
            # For withholding, prefer employee['withholding_rate'] but fall back to employee['tax_rate'] (some CSVs store it there).
            display_w_rate = float(employee.get('withholding_rate', 0) or employee.get('tax_rate', 0) or 0)
            display_p_rate = float(employee.get('p_tax_rate', 0) or 0)

            def label_with_rate(kind, rate):
                if rate and rate > 0:
                    return f"{kind} {rate*100:.0f}%"
                return f"{kind}"

            w_label_15 = label_with_rate('Withholding Tax (15th)', display_w_rate)
            p_label_15 = label_with_rate('Percentage Tax (15th)', display_p_rate)
            w_label_30 = label_with_rate('Withholding Tax (30th)', display_w_rate)
            p_label_30 = label_with_rate('Percentage Tax (30th)', display_p_rate)

            summary_data = [
                ['Description', 'Amount']
            ]
            summary_data += [
                ['Gross Pay', f"{gross:.2f}"]
            ]
            summary_data += [
                [w_label_15, f"{w_15:.2f}"],
                [p_label_15, f"{p_15:.2f}"],
                ['Total (15th)', f"{total_15:.2f}"],
                [w_label_30, f"{w_30:.2f}"],
                [p_label_30, f"{p_30:.2f}"],
                ['Total (30th)', f"{total_30:.2f}"],
                ['Total Deductions', f"{total_deductions:.2f}"],
                ['NET PAY', f"{net_pay:.2f}"]
            ]

            summary_table = Table(summary_data, colWidths=[260, 120])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(summary_table)
            story.append(Spacer(1, 20))

            # Payment Information: intentionally omitted per user request
            # (Payment Method / Bank / Account lines removed)

            # Footer
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1
            )

            story.append(Paragraph("For questions about this payslip, contact the University Accounting Office", footer_style))
            story.append(Paragraph("This payslip serves as official record of payment", footer_style))

            # Add signature line
            story.append(Spacer(1, 30))
            story.append(Paragraph("______________________________", emp_info_style))
            story.append(Paragraph("Employee Signature", emp_info_style))

            # Build PDF
            doc.build(story)
            return True

        except Exception as e:
            print(f"Error generating PDF for {employee['name']}: {e}")
            return False
    def _excel_find_sheet(self, wb):
        # Prefer the exact name, else fallback to the first sheet
        if 'Payslip Template' in wb.sheetnames:
            return wb['Payslip Template']
        return wb[wb.sheetnames[0]]
    def _excel_safe_set(self, ws, cell, value):
        try:
            ws[cell] = value
        except Exception:
            pass
    def _excel_fill_template(self, ws, employee, placement='both'):
        """
        Fill the Excel payslip to match your sheet and support up to 5 periods.
        Overwrites values (no formulas needed).
        """

        # ========= CELL MAP (from your screenshots) =========
        # Left summary
        CELL_EMP_NAME      = 'D9'
        CELL_RATE          = 'D11'
        CELL_TOTAL_HOURS   = 'D12'
        CELL_SAL_EARNED    = 'D13'  # Salary Earned
        CELL_ADJUSTMENT    = 'D14'  # Adjustment: (Add'l TAX)
        CELL_SUB_TOTAL     = 'D15'  # Salary Earned - Adjustment
        CELL_LESS_TAX      = 'D16'  # SAME as Total Deduction
        CELL_NET_PAY       = 'D17'  # total net

        # Pay per period (below NET PAY)
        PAY_LABEL_COL      = 'B'
        PAY_AMOUNT_COL     = 'D'
        PAY_START_ROW      = 18
        PAY_MAX_ROWS       = 5

        # BREAKDOWN (DATE / HOURS) block
        DATE_COL           = 'E'
        HOURS_COL          = 'F'
        BR_START_ROW       = 12
        BR_MAX_ROWS        = 5

        # Deduction amounts
        CELL_W_15_AMT      = 'I11'
        CELL_P_15_AMT      = 'I12'
        CELL_15_TOTAL      = 'I13'
        CELL_W_30_AMT      = 'I14'
        CELL_P_30_AMT      = 'I15'
        CELL_30_TOTAL      = 'I16'
        CELL_TOT_DED       = 'I17'

        # Deduction % (rates) – what you asked to fill
        CELL_W_15_RATE     = 'H11'
        CELL_P_15_RATE     = 'H12'
        CELL_W_30_RATE     = 'H14'
        CELL_P_30_RATE     = 'H15'

        # Optional header text of periods
        CELL_HEADER_PERIODS = 'H8'

        # ========= Gather data =========
        name        = employee.get('name', '')
        rate        = float(employee.get('hourly_rate', 0) or 0)
        adj_amt     = float(employee.get('adjustment_amount', 0) or 0)

        # Keep CSV period order
        period_entries = []
        for p in getattr(self, 'periods', []):
            lbl = p.get('name')
            rec = employee.get('periods', {}).get(lbl)
            if not rec:
                continue
            try:
                hrs = float(rec.get('hours', 0) or 0)
            except Exception:
                hrs = 0.0
            period_entries.append((lbl, hrs))
        if not period_entries:
            period_entries = [('', 0.0)]

        total_hours = sum(h for _, h in period_entries)
        sal_earned  = round(rate * total_hours, 2)
        sub_total   = round(sal_earned - adj_amt, 2)

        # Deductions: prefer absolute amounts, otherwise compute from rates
        gross_for_tax = sal_earned
        w_rate = float(employee.get('withholding_rate', 0) or employee.get('tax_rate', 0) or 0)
        p_rate = float(employee.get('p_tax_rate', 0) or 0)

        w_amt  = float(employee.get('withholding_amount', 0) or 0)
        if w_amt == 0 and w_rate:
            w_amt = round(gross_for_tax * w_rate, 2)

        p_amt  = float(employee.get('percent_tax', 0) or 0)
        if p_amt == 0 and p_rate:
            p_amt = round(gross_for_tax * p_rate, 2)

        # If rates missing but amounts present, back-calc rates so H11/H12/H14/H15 show %
        if (not w_rate) and w_amt and gross_for_tax:
            w_rate = w_amt / gross_for_tax
        if (not p_rate) and p_amt and gross_for_tax:
            p_rate = p_amt / gross_for_tax

        place = (placement or 'both').lower()
        def split_amount(v):
            if place == '15':  return v, 0.0
            if place == '30':  return 0.0, v
            return v/2.0, v/2.0

        w_15, w_30 = split_amount(w_amt)
        p_15, p_30 = split_amount(p_amt)

        total_15  = round(w_15 + p_15, 2)
        total_30  = round(w_30 + p_30, 2)
        total_ded = round(total_15 + total_30, 2)
        net_pay   = round(sub_total - total_ded, 2)

        # Helper to set a percent (decimal → show as "10%")
        def _set_percent(addr, val):
            try:
                c = ws[addr]
                if val is None or val == 0:
                    c.value = ""
                else:
                    c.value = float(val)   # 0.10
                    c.number_format = '0%' # or '0.00%' if you prefer two decimals
            except Exception:
                pass

        # ========= Left summary =========
        try: ws[CELL_EMP_NAME]    = name
        except Exception: pass
        try: ws[CELL_RATE]        = rate
        except Exception: pass
        try: ws[CELL_TOTAL_HOURS] = total_hours
        except Exception: pass
        try: ws[CELL_SAL_EARNED]  = sal_earned
        except Exception: pass
        try: ws[CELL_ADJUSTMENT]  = adj_amt
        except Exception: pass
        try: ws[CELL_SUB_TOTAL]   = sub_total
        except Exception: pass
        try: ws[CELL_LESS_TAX]    = total_ded
        except Exception: pass
        try: ws[CELL_NET_PAY]     = net_pay
        except Exception: pass

        try:
            ws[CELL_HEADER_PERIODS] = ', '.join(lbl for lbl, _ in period_entries if lbl) or ws[CELL_HEADER_PERIODS].value
        except Exception:
            pass

        # ========= BREAKDOWN (DATE/HOURS) =========
        for i in range(BR_MAX_ROWS):
            r = BR_START_ROW + i
            try:
                ws[f"{DATE_COL}{r}"]  = ""
                ws[f"{HOURS_COL}{r}"] = ""
            except Exception: pass
        for i, (lbl, hrs) in enumerate(period_entries[:BR_MAX_ROWS]):
            r = BR_START_ROW + i
            try:
                ws[f"{DATE_COL}{r}"]  = lbl
                ws[f"{HOURS_COL}{r}"] = hrs
            except Exception: pass

        # ========= Pay per period (below NET PAY) =========
        for i in range(PAY_MAX_ROWS):
            r = PAY_START_ROW + i
            try:
                ws[f"{PAY_LABEL_COL}{r}"]  = ""
                ws[f"{PAY_AMOUNT_COL}{r}"] = ""
            except Exception: pass
        for i, (lbl, hrs) in enumerate(period_entries[:PAY_MAX_ROWS]):
            r = PAY_START_ROW + i
            amt = round(hrs * rate, 2)
            try:
                ws[f"{PAY_LABEL_COL}{r}"]  = lbl
                ws[f"{PAY_AMOUNT_COL}{r}"] = amt
            except Exception: pass

        if len(period_entries) > PAY_MAX_ROWS:
            last_r = PAY_START_ROW + PAY_MAX_ROWS - 1
            extra_labels = [lbl for lbl, _ in period_entries[PAY_MAX_ROWS:]]
            extra_amt    = round(sum(hrs*rate for _, hrs in period_entries[PAY_MAX_ROWS:]), 2)
            try:
                cur_lbl = str(ws[f"{PAY_LABEL_COL}{last_r}"].value or "").strip()
                ws[f"{PAY_LABEL_COL}{last_r}"] = (cur_lbl + "; " if cur_lbl else "") + "; ".join(extra_labels)
                cur_amt = float(ws[f"{PAY_AMOUNT_COL}{last_r}"].value or 0)
                ws[f"{PAY_AMOUNT_COL}{last_r}"] = round(cur_amt + extra_amt, 2)
            except Exception:
                pass

        # ========= DEDUCTION RATES (H11/H12/H14/H15) =========
        _set_percent(CELL_W_15_RATE, w_rate)
        _set_percent(CELL_P_15_RATE, p_rate)
        _set_percent(CELL_W_30_RATE, w_rate)
        _set_percent(CELL_P_30_RATE, p_rate)

        # ========= DEDUCTION AMOUNTS (I11–I17) =========
        try: ws[CELL_W_15_AMT] = round(w_15, 2)
        except Exception: pass
        try: ws[CELL_P_15_AMT] = round(p_15, 2)
        except Exception: pass
        try: ws[CELL_15_TOTAL] = total_15
        except Exception: pass
        try: ws[CELL_W_30_AMT] = round(w_30, 2)
        except Exception: pass
        try: ws[CELL_P_30_AMT] = round(p_30, 2)
        except Exception: pass
        try: ws[CELL_30_TOTAL] = total_30
        except Exception: pass
        try: ws[CELL_TOT_DED]  = total_ded
        except Exception: pass
        # --- lock to a single tight page ---
        # --- lock to a single tight page (with safety buffer to avoid clipping) ---
        TOP_LEFT_COL = 'B'
        TOP_LEFT_ROW = 7

        last_breakdown_row = BR_START_ROW + min(len(period_entries), BR_MAX_ROWS) - 1
        last_pay_row       = PAY_START_ROW + min(len(period_entries), PAY_MAX_ROWS) - 1

        # Include the footer rows and add a tiny padding row
        STATIC_FOOTER_MIN_ROW = 22     # was 21; include the "Date:" row fully
        EXTRA_PADDING_ROWS    = 1      # 1 extra row avoids bottom clipping
        RIGHTMOST_COL         = 'J'    # extend one column past I to avoid right-edge clipping

        bottom_row = max(last_breakdown_row, last_pay_row, STATIC_FOOTER_MIN_ROW) + EXTRA_PADDING_ROWS
        print_area  = f"{TOP_LEFT_COL}{TOP_LEFT_ROW}:{RIGHTMOST_COL}{bottom_row}"

        self._apply_single_page_print(ws, print_area)


    def _convert_xlsx_to_pdf(self, xlsx_path, pdf_path):
        """
        Use LibreOffice to convert XLSX -> PDF. Returns True on success.
        """
        soffice = shutil.which('soffice') or shutil.which('libreoffice')
        if not soffice:
            print("LibreOffice not found; cannot export template to PDF.")
            return False

        outdir = os.path.dirname(os.path.abspath(pdf_path))
        os.makedirs(outdir, exist_ok=True)

        # Convert; LibreOffice writes PDF next to file using the same basename.
        cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", outdir, xlsx_path]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except subprocess.CalledProcessError as e:
            print("LibreOffice conversion failed:", e)
            return False

        # LibreOffice will produce: outdir / basename(xlsx).pdf
        base = os.path.splitext(os.path.basename(xlsx_path))[0]
        produced = os.path.join(outdir, base + ".pdf")

        # If the produced file isn't the desired name, rename it
        if os.path.abspath(produced) != os.path.abspath(pdf_path) and os.path.exists(produced):
            try:
                os.replace(produced, pdf_path)
            except Exception:
                # copy/remove fallback
                import shutil as _sh
                _sh.copyfile(produced, pdf_path)
                try:
                    os.remove(produced)
                except Exception:
                    pass

        return os.path.exists(pdf_path)
    def generate_pdf_from_excel_template(self, employee, output_filename, withholding_placement='both'):
        """
        Fill 'Template Payslip.xlsx' and export to a PDF that matches the Excel template exactly.
        Falls back to False if openpyxl or the template is unavailable.
        """
        if not HAS_OPENPYXL or not os.path.exists(self.template_path):
            return False

        try:
            wb = load_workbook(self.template_path, data_only=False)
            ws = self._excel_find_sheet(wb)

            # Fill the template cells
            self._excel_fill_template(ws, employee, withholding_placement)

            # Save a per-employee XLSX in a temp folder
            tmp_dir = os.path.join("output_pdfs", "tmp")
            os.makedirs(tmp_dir, exist_ok=True)

            seq_num = int(employee.get('seq')) if str(employee.get('seq', '')).isdigit() else 0
            clean_name = "".join(c for c in employee.get('name', '') if c.isalnum() or c in (' ', '-', '_')).rstrip()
            xlsx_temp = os.path.join(tmp_dir, f"payslip_{seq_num:03d}_{clean_name.replace(' ', '_')}.xlsx")
            wb.save(xlsx_temp)

            # Convert to PDF via LibreOffice
            ok = self._convert_xlsx_to_pdf(xlsx_temp, output_filename)
            return ok
        except Exception as e:
            print(f"Template -> PDF failed for {employee.get('name','')}: {e}")
            return False

    def process_payroll_to_pdfs(self, csv_filename, withholding_placement='both'):
        """Main function to process CSV and generate PDFs"""
        print("=== Dynamic Payroll PDF Generator ===")
        print(f"Processing: {csv_filename}")
        print()

        # Step 1: Detect periods
        print("Step 1: Detecting pay periods...")
        if not self.detect_periods(csv_filename):
            print("Failed to detect periods. Please check CSV format.")
            return False

        # Step 2: Load employee data
        print("\nStep 2: Loading employee data...")
        if not self.load_employee_data(csv_filename):
            print("Failed to load employee data.")
            return False

        # Step 3: Generate PDFs
        print(f"\nStep 3: Generating {len(self.employees)} PDF payslips...")

        success_count = 0
        for i, employee in enumerate(self.employees, 1):
            # Create filename
            clean_name = "".join(c for c in employee['name'] if c.isalnum() or c in (' ', '-', '_')).rstrip()
            seq_num = int(employee['seq']) if employee['seq'].isdigit() else 0
            filename = f"payslip_{seq_num:03d}_{clean_name.replace(' ', '_')}.pdf"
            filepath = os.path.join("output_pdfs", filename)

            # Generate PDF
            if self.generate_pdf_payslip(employee, filepath, withholding_placement):
                success_count += 1
                print(f"  [{i}/{len(self.employees)}] Generated: {filename}")
            else:
                print(f"  [{i}/{len(self.employees)}] Failed: {employee['name']}")

        print("\n=== Processing Complete ===")
        print(f"Successfully generated {success_count}/{len(self.employees)} PDF payslips")
        print(f"Output location: {os.path.abspath('output_pdfs')}")
        return success_count > 0

    # ----------------- Convenience: send all payslips -----------------
    def send_all_payslips(self, output_dir='output_pdfs', subject_tpl=None, body_tpl=None, throttle_seconds=0.5, dry_run=True, withholding_placement='both', progress_cb=None, force_regen=False):
        """
        Generate PDFs (into output_dir) and send them via Mailer.send.
        Returns a list of dicts per employee: {name, seq, status, error, path}
        progress_cb, if provided, is called with a single string argument for progress messages.
        """
        try:
            from mailer import Mailer
        except Exception:
            Mailer = None

        os.makedirs(output_dir, exist_ok=True)
        results = []

        mailer = Mailer() if Mailer is not None else None
        total = len(self.employees)

        if getattr(self, 'email_store', None):
            try:
                self.email_store.apply_to_employees(self.employees)
            except Exception:
                pass

        # If several employees are missing emails, try to merge addresses from a local emails.csv
        missing_any = any(not (emp.get('email') or emp.get('personal_email') or emp.get('work_email')) for emp in self.employees)
        email_map = {}
        if missing_any:
            emails_csv = os.path.join(os.getcwd(), 'emails.csv')
            if os.path.exists(emails_csv):
                try:
                    if callable(progress_cb):
                        progress_cb(f"Loading emails from {emails_csv}")
                    with open(emails_csv, 'r', encoding='utf-8', newline='') as fh:
                        import csv as _csv
                        rdr = _csv.DictReader(fh)
                        for r in rdr:
                            if r is None:
                                continue
                            # prefer seq as key, fallback to account_no or name
                            if r.get('seq'):
                                key = str(r.get('seq')).strip()
                                email_map.setdefault('seq', {})[key] = r.get('email')
                            acc = (r.get('account_no') or '').strip()
                            if acc:
                                email_map.setdefault('account_no', {})[acc] = r.get('email')
                            nm = (r.get('name') or '').strip().lower()
                            if nm:
                                email_map.setdefault('name', {})[nm] = r.get('email')
                except Exception:
                    # ignore errors reading external email file
                    email_map = {}

        for idx, emp in enumerate(self.employees, start=1):
            clean_name = "".join(c for c in emp.get('name', '') if c.isalnum() or c in (' ', '-', '_')).rstrip()
            seq_num = int(emp.get('seq')) if str(emp.get('seq', '')).isdigit() else 0
            filename = f"payslip_{seq_num:03d}_{clean_name.replace(' ', '_')}.pdf"
            outpath = os.path.join(output_dir, filename)

            # Log progress per employee
            msg_line = f"[{idx}/{total}] Generating: {filename}"
            if callable(progress_cb):
                try:
                    progress_cb(msg_line)
                except Exception:
                    pass
            else:
                print(msg_line)

            # If PDF already exists and not forcing regeneration, skip generation
            ok = False
            try:
                if (not force_regen) and os.path.exists(outpath) and os.path.getsize(outpath) > 0:
                    ok = True
                    if callable(progress_cb):
                        try:
                            progress_cb(f"[{idx}/{total}] Skipping regen, existing PDF found: {filename}")
                        except Exception:
                            pass
                else:
                    ok = self.generate_pdf_payslip(emp, outpath, withholding_placement)
            except Exception as e:
                ok = False
                if callable(progress_cb):
                    try:
                        progress_cb(f"[{idx}/{total}] Error generating {filename}: {e}")
                    except Exception:
                        pass
            if not ok:
                results.append({"name": emp.get('name', ''), "seq": emp.get('seq', ''), "status": "skipped:no_pdf", "error": "pdf_failed", "path": outpath, "email": None})
                continue

            # Merge-in email from emails.csv (if present) when missing
            to_email = emp.get('email') or emp.get('personal_email') or emp.get('work_email') or None
            if not to_email and email_map:
                # try seq
                seq_key = str(emp.get('seq', '')).strip()
                if seq_key and email_map.get('seq', {}).get(seq_key):
                    to_email = email_map['seq'][seq_key]
                # try account_no
                if not to_email and emp.get('account_no') and email_map.get('account_no', {}).get(emp.get('account_no')):
                    to_email = email_map['account_no'][emp.get('account_no')]
                # try name
                if not to_email:
                    nm = (emp.get('name') or '').strip().lower()
                    if nm and email_map.get('name', {}).get(nm):
                        to_email = email_map['name'][nm]

            # If we found a merged email, attach back into employee for future runs
            if to_email:
                try:
                    if not emp.get('email'):
                        emp['email'] = to_email
                except Exception:
                    pass
                try:
                    self.remember_employee_email(emp, to_email)
                except Exception:
                    pass
            if not to_email:
                results.append({"name": emp.get('name', ''), "seq": emp.get('seq', ''), "status": "skipped:no_email", "error": "no_recipient", "path": outpath, "email": None})
                continue

            if dry_run or mailer is None:
                results.append({"name": emp.get('name', ''), "seq": emp.get('seq', ''), "status": "dry_run", "error": None, "path": outpath, "email": to_email})
            else:
                subject = (subject_tpl or "Payslip for {name}").format(name=emp.get('name', ''))
                body = (body_tpl or "Please find attached your payslip.").format(name=emp.get('name', ''))
                try:
                    mailer.send(to_email, subject, body, attachments=[outpath])
                    results.append({"name": emp.get('name', ''), "seq": emp.get('seq', ''), "status": "sent", "error": None, "path": outpath, "email": to_email})
                except Exception as e:
                    results.append({"name": emp.get('name', ''), "seq": emp.get('seq', ''), "status": "error", "error": str(e), "path": outpath, "email": to_email})

            # throttle between sends
            try:
                time.sleep(float(throttle_seconds or 0))
            except Exception:
                pass

        if getattr(self, 'email_store', None):
            try:
                self.email_store.export_to_csv(os.path.join(os.getcwd(), 'emails.csv'))
            except Exception:
                pass

        return results
    
    def _apply_single_page_print(self, ws, print_area):
        """
        Constrain export to a single page exactly covering `print_area`.
        """
        try:
            # Only print the block we filled
            ws.print_area = print_area
            ws.print_title_rows = ''
            ws.print_title_cols = ''
        except Exception:
            pass

        # Clear any manual page breaks that can create a blank Page 2
        try:
            if PageBreak is not None:
                ws.row_breaks = PageBreak()
                ws.col_breaks = PageBreak()
            else:
                # fallback for older openpyxl (<=3.0.x)
                from openpyxl.worksheet.pagebreak import RowBreaks, ColBreaks  # type: ignore
                ws.row_breaks = RowBreaks()
                ws.col_breaks = ColBreaks()
        except Exception:
            pass


        # Fit to 1 page
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered   = False

        # Paper & margins (a touch wider to prevent cuts)
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize   = 9   # A4 (use 11 for A5 if you prefer smaller paper)

        from openpyxl.worksheet.page import PageMargins
        ws.page_margins = PageMargins(
            left=0.3, right=0.3, top=0.35, bottom=0.35, header=0.1, footer=0.1
        )

        try:
            ws.page_margins = PageMargins(
                left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1
            )
        except Exception:
            pass
    def _locate_soffice(self):
        """Locate a soffice/libreoffice binary if present (PATH or typical installs)."""
        candidates = [
            shutil.which('soffice'),
            shutil.which('libreoffice'),
        ]
        if platform.system() == 'Windows':
            candidates += [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]
        for c in candidates:
            if c and os.path.exists(c):
                return c
        return None

    def _convert_with_soffice(self, soffice, xlsx_path, pdf_path):
        """Convert using LibreOffice CLI (soffice). Returns True on success."""
        try:
            outdir = os.path.dirname(os.path.abspath(pdf_path)) or "."
            os.makedirs(outdir, exist_ok=True)
            cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", outdir, xlsx_path]
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            produced = os.path.join(outdir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
            if os.path.exists(produced):
                # if produced name differs from desired, move/replace
                if os.path.abspath(produced) != os.path.abspath(pdf_path):
                    try:
                        os.replace(produced, pdf_path)
                    except Exception:
                        shutil.copyfile(produced, pdf_path)
                        try: os.remove(produced)
                        except Exception: pass
            return os.path.exists(pdf_path)
        except Exception as e:
            print(f"[soffice] conversion failed: {e}")
            return False

    def _convert_with_excel(self, xlsx_path, pdf_path, sheet_name=None, print_area=None):
        """
        Convert using Excel COM automation (Windows only). Returns True on success.
        Requires pywin32 (win32com). Opens Excel hidden, exports sheet as PDF, quits.
        """
        if platform.system() != 'Windows':
            print("[excel] Excel COM only available on Windows.")
            return False

        try:
            import pythoncom
            import win32com.client as win32
        except Exception as e:
            print("[excel] pywin32 not available:", e)
            return False

        xlsx_path = os.path.abspath(xlsx_path)
        pdf_path = os.path.abspath(pdf_path)
        outdir = os.path.dirname(pdf_path)
        os.makedirs(outdir, exist_ok=True)

        excel = None
        wb = None
        try:
            pythoncom.CoInitialize()
            # Use DispatchEx to ensure a new Excel instance (avoids interfering with user Excel)
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # Open workbook (not read-only to allow PrintArea to be set if requested)
            wb = excel.Workbooks.Open(xlsx_path, ReadOnly=False)

            # Select sheet
            ws = None
            if sheet_name:
                try:
                    ws = wb.Worksheets(sheet_name)
                except Exception:
                    ws = None
            if ws is None:
                ws = wb.Worksheets(1)

            # If a print area is provided, set it so Export uses the right bounds
            if print_area:
                try:
                    ws.PageSetup.PrintArea = str(print_area)
                except Exception as e:
                    print("[excel] failed to set PrintArea:", e)

            # Export to PDF. 0 = xlTypePDF
            xlTypePDF = 0
            # Export the worksheet as a PDF file
            ws.ExportAsFixedFormat(
                Type=xlTypePDF,
                Filename=pdf_path,
                Quality=0,                 # 0 = standard
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )

            # small delay to ensure file is written
            time.sleep(0.1)
            return os.path.exists(pdf_path)
        except Exception as e:
            print("[excel] Export failed:", e)
            return False
        finally:
            try:
                if wb is not None:
                    # Close workbook without saving changes
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _convert_xlsx_to_pdf(self, xlsx_path, pdf_path):
        """
        Unified converter: on Windows use Excel COM; on other OS try LibreOffice then Excel fallback.
        Returns True on success.
        """
        xlsx_path = os.path.abspath(xlsx_path)
        pdf_path = os.path.abspath(pdf_path)

        # On Windows prefer Excel COM (native)
        if platform.system() == 'Windows':
            print("[convert] Using Excel COM on Windows...")
            ok = self._convert_with_excel(xlsx_path, pdf_path, sheet_name=getattr(self, 'template_sheet_name', None))
            if ok:
                return True
            # if Excel fails, try soffice as a last resort
            soffice = self._locate_soffice()
            if soffice:
                print("[convert] Excel failed — trying LibreOffice (soffice)...")
                return self._convert_with_soffice(soffice, xlsx_path, pdf_path)
            return False

        # On Linux/macOS: try soffice first
        soffice = self._locate_soffice()
        if soffice:
            print(f"[convert] Found soffice at {soffice}; using it.")
            ok = self._convert_with_soffice(soffice, xlsx_path, pdf_path)
            if ok:
                return True
            print("[convert] soffice failed; trying Excel COM (pywin32) fallback if available.")
            # try excel if pywin32 is present (rare on non-Windows)
            try:
                import win32com.client  # type: ignore
                ok2 = self._convert_with_excel(xlsx_path, pdf_path, sheet_name=getattr(self, 'template_sheet_name', None))
                return ok2
            except Exception:
                return False

        # Neither soffice nor excel available
        print("[convert] No PDF converter found (soffice or Excel COM).")
        return False
        
def main():
    """Main function"""
    processor = DynamicPayrollProcessor()

    # Check for CSV files
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv') and f != 'Template Payslip.csv']

    if not csv_files:
        print("No CSV files found in current directory.")
        csv_file = input("Enter CSV filename: ").strip()
    else:
        print("Found CSV files:")
        for i, file in enumerate(csv_files, 1):
            print(f"{i}. {file}")

        choice = input(f"Select CSV file (1-{len(csv_files)}) or enter filename: ").strip()

        if choice.isdigit() and 1 <= int(choice) <= len(csv_files):
            csv_file = csv_files[int(choice) - 1]
        else:
            csv_file = choice

    # Ask user where to place withholding amounts: 15, 30, or both
    placement = input("Withholding placement? (15/30/both) [both]: ").strip().lower() or 'both'
    if placement not in ('15', '30', 'both'):
        placement = 'both'

    if not os.path.exists(csv_file):
        print(f"File '{csv_file}' not found.")
        return

    # Process the payroll (pass withholding placement)
    processor.process_payroll_to_pdfs(csv_file, withholding_placement=placement)

if __name__ == "__main__":
    main()